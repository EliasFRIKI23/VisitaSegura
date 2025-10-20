import pandas as pd
from datetime import datetime
from typing import List, Dict
import os

class ExcelExporter:
    """Clase para exportar datos de visitantes a Excel"""
    
    def __init__(self):
        self.report_dir = "report"
        self._ensure_report_directory()
    
    def _ensure_report_directory(self):
        """Asegura que el directorio de reportes existe"""
        if not os.path.exists(self.report_dir):
            os.makedirs(self.report_dir)
    
    def export_visitors_report(self, visitors_data: List[Dict], filename: str = None) -> str:
        """
        Exporta los datos de visitantes actuales a un archivo Excel
        
        Args:
            visitors_data: Lista de diccionarios con datos de visitantes
            filename: Nombre del archivo (opcional, se genera automáticamente si no se proporciona)
        
        Returns:
            str: Ruta del archivo generado
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_visitantes_actuales_{timestamp}.xlsx"
        
        # Asegurar que el archivo tenga extensión .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.report_dir, filename)
        
        try:
            # Crear DataFrame con los datos
            df = pd.DataFrame(visitors_data)
            
            # Asegurar que tenemos todas las columnas necesarias
            expected_columns = [
                'nombre', 'rut', 'fecha_entrada', 'fecha_salida', 
                'destino', 'acompañante', 'estado_visita'
            ]
            
            # Agregar usuario_registrador si no existe (compatibilidad hacia atrás)
            if 'usuario_registrador' not in df.columns:
                df['usuario_registrador'] = 'Sistema'
            
            # Filtrar solo las columnas que necesitamos
            df = df[expected_columns + ['usuario_registrador']]
            
            # Renombrar columnas para mejor presentación
            df.columns = [
                'Nombre del Visitante',
                'RUT',
                'Fecha y Hora de Entrada',
                'Fecha y Hora de Salida',
                'Destino/Lugar',
                'Acompañante',
                'Estado de Visita',
                'Registrado por'
            ]
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Visitantes Actuales', index=False)
                
                # Obtener el workbook y worksheet para aplicar formato
                workbook = writer.book
                worksheet = writer.sheets['Visitantes Actuales']
                
                # Aplicar formato a las columnas
                self._format_worksheet(worksheet, len(visitors_data))
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error al exportar a Excel: {str(e)}")
    
    def _format_worksheet(self, worksheet, num_rows):
        """Aplica formato al worksheet de Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Estilo para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para las celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formato al encabezado (fila 1)
        for col in range(1, 9):  # 8 columnas (incluyendo usuario registrador)
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Aplicar formato a las celdas de datos
        for row in range(2, num_rows + 2):  # +2 porque empezamos desde la fila 2
            for col in range(1, 9):  # 8 columnas
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [25, 15, 20, 20, 20, 20, 15, 18]  # Anchos para cada columna (incluyendo usuario registrador)
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = width
    
    def get_report_info(self, visitors_data: List[Dict]) -> Dict:
        """Obtiene información del reporte para mostrar en la interfaz"""
        total_visitors = len(visitors_data)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Contar visitantes por destino
        destinations = {}
        for visitor in visitors_data:
            dest = visitor['destino']
            destinations[dest] = destinations.get(dest, 0) + 1
        
        # Contar visitantes por usuario registrador
        registradores = {}
        for visitor in visitors_data:
            registrador = visitor.get('usuario_registrador', 'Sistema')
            registradores[registrador] = registradores.get(registrador, 0) + 1
        
        return {
            'total_visitors': total_visitors,
            'generation_time': current_time,
            'destinations': destinations,
            'registradores': registradores
        }

