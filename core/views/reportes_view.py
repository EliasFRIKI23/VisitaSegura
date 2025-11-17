from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
                               QFrame, QGridLayout, QTableWidget, QTableWidgetItem, 
                               QHeaderView, QMessageBox, QFileDialog, QProgressBar, QComboBox,
                               QSizePolicy, QScrollArea, QAbstractItemView)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QFont, QPixmap, QColor, QGuiApplication
from core.ui.icon_loader import get_icon_for_emoji
import sys
import os
from datetime import datetime, timedelta
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
try:
    from core.theme import (
        DUOC_PRIMARY, DUOC_SECONDARY, DUOC_SUCCESS, DUOC_DANGER, DUOC_INFO,
        darken_color as duoc_darken, lighten_color as duoc_lighten, get_standard_button_style,
        format_rut_display, get_current_user
    )
    from core.ui import configure_modern_table, apply_modern_table_theme
except Exception:
    DUOC_PRIMARY = "#003A70"
    DUOC_SECONDARY = "#FFB81C"
    DUOC_SUCCESS = "#28a745"
    DUOC_DANGER = "#dc3545"
    DUOC_INFO = "#17a2b8"
    def duoc_darken(color, factor=0.2):
        color = color.lstrip('#')
        r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        r = max(0, int(r * (1 - factor)))
        g = max(0, int(g * (1 - factor)))
        b = max(0, int(b * (1 - factor)))
        return f"#{r:02x}{g:02x}{b:02x}"
    def duoc_lighten(color, factor=0.1):
        color = color.lstrip('#')
        r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        r = min(255, int(r + (255 - r) * factor))
        g = min(255, int(g + (255 - g) * factor))
        b = min(255, int(b + (255 - b) * factor))
        return f"#{r:02x}{g:02x}{b:02x}"
    def get_standard_button_style(color, text_color=None):
        resolved_color = text_color or ('#000000' if color in [DUOC_SECONDARY, "#ffc107"] else '#ffffff')
        return f"""
            QPushButton {{
                background-color: {color};
                color: {resolved_color};
                border: none;
                border-radius: 6px;
                padding: 10px 16px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {duoc_darken(color, 0.1)};
            }}
            QPushButton:disabled {{
                background-color: #6c757d;
                color: #adb5bd;
            }}
        """
    def format_rut_display(rut):
        """Fallback si no se puede importar la función"""
        return rut
    def get_current_user():
        """Fallback si no se puede importar la función"""
        return "Sistema"
    def configure_modern_table(table, **kwargs):  # type: ignore
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setShowGrid(False)
        table.verticalHeader().setVisible(False)
    def apply_modern_table_theme(table, dark_mode=False):  # type: ignore
        base_bg = "#0f172a" if dark_mode else "#ffffff"
        base_fg = "#e2e8f0" if dark_mode else "#1f2937"
        table.setStyleSheet(
            f"""
            QTableWidget {{
                background-color: {base_bg};
                color: {base_fg};
                border: none;
            }}
            """
        )
import numpy as np

# Agregar el directorio padre al path para importar módulos
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from core.visitors import VisitorManager
    from core.excel_exporter import ExcelExporter
    from core.auth_manager import AuthManager
except ImportError:
    from visitors import VisitorManager
    from excel_exporter import ExcelExporter
    from auth_manager import AuthManager

class ChartWidget(QWidget):
    """Widget personalizado para mostrar gráficos"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        try:
            # Configuración optimizada para PySide6 y layout horizontal con scroll
            self.figure = Figure(figsize=(8, 6), dpi=110)
            self.canvas = FigureCanvas(self.figure)
            
            # Layout simple
            layout = QVBoxLayout(self)
            layout.setContentsMargins(5, 5, 5, 5)
            layout.addWidget(self.canvas)
            
            # Tamaño optimizado para layout horizontal sin scroll horizontal
            self.setFixedSize(450, 350)
            
        except Exception as e:
            print(f"Error al inicializar ChartWidget: {e}")
    
    def _style_axes(self, ax, facecolor="#ffffff"):
        """Aplica un estilo limpio y legible a los ejes del gráfico."""
        try:
            ax.set_facecolor(facecolor)
            # Limpiar spines y ticks
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color('#cccccc')
            ax.spines['bottom'].set_color('#cccccc')
            ax.tick_params(colors='#000000', labelsize=9)  # Texto negro
            # Grid sutil en eje Y
            ax.grid(True, alpha=0.25, color='#d9d9d9', linestyle='--', linewidth=0.7, axis='y')
        except Exception:
            pass
    
    def clear(self):
        """Limpia el gráfico"""
        self.figure.clear()
    
    def plot_visitors_by_day(self, data):
        """Gráfico de visitantes por día"""
        try:
            self.clear()
            ax = self.figure.add_subplot(111)
            self._style_axes(ax)
            
            if not data:
                ax.text(0.5, 0.5, 'Sin datos disponibles', ha='center', va='center', 
                       transform=ax.transAxes, fontsize=12)
                self.canvas.draw()
                return
            
            dates = list(data.keys())
            counts = list(data.values())
            
            # Gráfico de barras con colores (variación suave según valor)
            max_count = max(counts) if counts else 1
            colors = [duoc_lighten(DUOC_PRIMARY, min(0.35, 0.1 + (c / max_count) * 0.25)) for c in counts]
            bars = ax.bar(range(len(dates)), counts, color=colors, alpha=0.95, edgecolor='white', linewidth=0.8)
            
            # Agregar valores en las barras
            for i, count in enumerate(counts):
                if count > 0:
                    ax.text(
                        i, count + 0.1, str(count), ha='center', va='bottom', fontweight='bold', fontsize=9,
                        color='#2c3e50', bbox=dict(boxstyle='round,pad=0.2', facecolor='#f8f9fa', edgecolor='none', alpha=0.9)
                    )
            
            ax.set_title('Visitantes por Día (Últimos 7 días)', fontsize=12, fontweight='bold', pad=15, color='#000000')
            ax.set_ylabel('Número de Visitantes', fontsize=10, color='#000000')
            ax.set_xlabel('Fecha', fontsize=10, color='#000000')
            ax.set_xticks(range(len(dates)))
            ax.set_xticklabels([d.strftime('%d/%m') for d in dates], rotation=45, fontsize=9)
            
            self.figure.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error al dibujar gráfico de visitantes por día: {e}")
    
    def plot_visitors_status(self, current, departed):
        """Gráfico de visitantes actuales vs que se fueron"""
        try:
            self.clear()
            ax = self.figure.add_subplot(111)
            
            categories = ['Visitantes Actuales', 'Visitantes que se Fueron']
            values = [current, departed]
            colors = [DUOC_SECONDARY, '#dc3545']

            bars = ax.bar(categories, values, color=colors)

            # Etiquetas de valores simples sobre cada barra
            for bar, value in zip(bars, values):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height, str(value),
                        ha='center', va='bottom', fontweight='bold')

            ax.set_title('Estado de Visitantes', fontsize=12, fontweight='bold', pad=12, color='#000000')
            ax.set_ylabel('Número de Visitantes', fontsize=10, color='#000000')
            ax.set_ylim(0, max(1, max(values)) * 1.2)
            plt.setp(ax.get_xticklabels(), rotation=0, ha='center')
            
            self.figure.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error al dibujar gráfico de estado: {e}")
    
    def plot_popular_destinations(self, destinations):
        """Gráfico de destinos más populares"""
        try:
            self.clear()
            ax = self.figure.add_subplot(111)
            self._style_axes(ax)
            
            if not destinations:
                ax.text(0.5, 0.5, 'Sin datos de destinos', ha='center', va='center', 
                       transform=ax.transAxes, fontsize=12)
                self.canvas.draw()
                return
                
            # Ordenar por cantidad
            sorted_dests = sorted(destinations.items(), key=lambda x: x[1], reverse=True)
            labels = [item[0] for item in sorted_dests]
            values = [item[1] for item in sorted_dests]
            
            # Gráfico de barras horizontales
            y_pos = np.arange(len(labels))
            bars = ax.barh(y_pos, values, color=duoc_lighten(DUOC_PRIMARY, 0.2), alpha=0.95, edgecolor='white', linewidth=1)
            
            # Agregar valores en las barras
            for i, (bar, value) in enumerate(zip(bars, values)):
                width = bar.get_width()
                ax.text(
                    width + 0.2, bar.get_y() + bar.get_height()/2.,
                    f'{value}', ha='left', va='center', fontweight='bold', fontsize=9,
                    color='#2c3e50', bbox=dict(boxstyle='round,pad=0.2', facecolor='#f8f9fa', edgecolor='none', alpha=0.9)
                )
            
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=10)
            ax.set_xlabel('Número de Visitas', fontsize=10, color='#000000')
            ax.set_title('Destinos Más Frecuentes', fontsize=12, fontweight='bold', pad=15, color='#000000')
            
            self.figure.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error al dibujar gráfico de destinos: {e}")

class ReportesView(QWidget):
    """Vista para reportes y estadísticas"""
    
    def __init__(self, parent=None, auth_manager=None):
        super().__init__(parent)
        self.visitor_manager = VisitorManager()
        self.excel_exporter = ExcelExporter()
        self.auth_manager = auth_manager or AuthManager()
        self.range_days = 7  # Rango inicial para el gráfico de visitantes por día
        
        # Configurar tamaño mínimo para mejor visualización de la tabla
        self.setMinimumSize(1300, 700)  # Ancho mínimo de 1300px para acomodar todas las columnas
        
        self.setup_ui()
        
        # Timer para actualizar automáticamente cada 10 segundos
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self.refresh_visitors_data)
        self.refresh_timer.start(10000)  # 10 segundos
        
        # Conectar el evento de cambio de tamaño para ajustar las columnas
        self.resizeEvent = self.on_resize_event
    
    def update_auth_manager(self, auth_manager):
        """Actualiza la instancia del AuthManager"""
        self.auth_manager = auth_manager
        self.update_export_button_visibility()
    
    def update_export_button_visibility(self):
        """Actualiza la visibilidad del botón de exportar según los permisos"""
        if hasattr(self, 'btn_export'):
            is_admin = self.auth_manager.is_admin()
            self.btn_export.setVisible(is_admin)
            self.btn_export.setEnabled(is_admin)
    
    def on_resize_event(self, event):
        """Maneja el evento de cambio de tamaño de la ventana"""
        super().resizeEvent(event)
        # Ajustar las columnas después de un pequeño delay para que el resize se complete
        QTimer.singleShot(100, self.adjust_table_columns)
    
    def adjust_table_columns(self):
        """Ajusta las columnas de la tabla según el tamaño actual"""
        if hasattr(self, 'visitors_table'):
            # Forzar el recálculo del layout de la tabla
            self.visitors_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.visitors_table.horizontalHeader().setStretchLastSection(True)
    
    def _get_screen_config(self):
        """Obtiene la configuración responsiva basada en el tamaño de pantalla"""
        available = QGuiApplication.primaryScreen().availableGeometry()
        if available.width() >= 1920:
            return {
                'margin': 30,
                'spacing': 25,
                'content_margin': 30,
                'content_spacing': 20,
                'btn_width': 180,
                'btn_height': 50,
                'btn_font_size': 16,
                'row_height': 45,
                'font_size': 12,
                'value_font_size': 28,
                'title_font_size': 14,
                'desc_font_size': 11,
                'border_radius': 15,
                'padding': 25
            }
        elif available.width() >= 1366:
            return {
                'margin': 20,
                'spacing': 20,
                'content_margin': 25,
                'content_spacing': 15,
                'btn_width': 150,
                'btn_height': 40,
                'btn_font_size': 14,
                'row_height': 40,
                'font_size': 11,
                'value_font_size': 24,
                'title_font_size': 12,
                'desc_font_size': 9,
                'border_radius': 12,
                'padding': 20
            }
        else:
            return {
                'margin': 15,
                'spacing': 15,
                'content_margin': 20,
                'content_spacing': 10,
                'btn_width': 120,
                'btn_height': 35,
                'btn_font_size': 12,
                'row_height': 35,
                'font_size': 10,
                'value_font_size': 20,
                'title_font_size': 10,
                'desc_font_size': 8,
                'border_radius': 10,
                'padding': 15
            }
    
    def setup_ui(self):
        """Configura la interfaz de usuario"""
        main_layout = QVBoxLayout(self)
        
        # Configuración responsiva centralizada
        self.screen_config = self._get_screen_config()
        
        main_layout.setContentsMargins(
            self.screen_config['margin'], 
            self.screen_config['margin'], 
            self.screen_config['margin'], 
            self.screen_config['margin']
        )
        main_layout.setSpacing(self.screen_config['spacing'])
        
        # Header con título y logo
        header_layout = QHBoxLayout()
        
        # Título
        title_layout = QVBoxLayout()
        title = QLabel("📊 Reportes y Estadísticas")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setAlignment(Qt.AlignLeft)
        
        subtitle = QLabel("Análisis y reportes del sistema de visitas")
        subtitle.setFont(QFont("Arial", 12))
        title.setAlignment(Qt.AlignLeft)
        
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        header_layout.addLayout(title_layout)
        
        header_layout.addStretch()
        
        # Logo Duoc
        logo_label = QLabel()
        logo_pixmap = QPixmap("Logo Duoc .png")
        if not logo_pixmap.isNull():
            scaled_pixmap = logo_pixmap.scaled(120, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
        else:
            logo_label.setText("🏢 Duoc UC")
            logo_font = QFont("Arial", 12, QFont.Bold)
            logo_label.setFont(logo_font)
        
        logo_label.setAlignment(Qt.AlignRight)
        header_layout.addWidget(logo_label)
        
        main_layout.addLayout(header_layout)
        
        # Crear área de scroll para todo el contenido
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # Sin scroll horizontal
        scroll_area.setFrameStyle(QFrame.NoFrame)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Widget contenedor para todo el contenido scrolleable
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        
        # Márgenes responsivos optimizados usando configuración centralizada
        content_layout.setContentsMargins(
            self.screen_config['content_margin'], 
            self.screen_config['content_margin'], 
            self.screen_config['content_margin'], 
            self.screen_config['content_margin']
        )
        content_layout.setSpacing(self.screen_config['content_spacing'])
        
        # Sección de Gráficos - Layout Horizontal
        # Encabezado de gráficos con selector de rango
        charts_header = QHBoxLayout()
        # Crear layout horizontal para título con icono
        charts_title_layout = QHBoxLayout()
        charts_title_layout.setContentsMargins(12, 0, 16, 0)
        charts_title_layout.setSpacing(8)
        
        charts_title_icon = get_icon_for_emoji("📊", 20)
        if not charts_title_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(charts_title_icon.pixmap(20, 20))
            charts_title_layout.addWidget(icon_label)
        
        charts_title = QLabel("Análisis Visual de Visitantes")
        charts_title.setFont(QFont("Arial", 18, QFont.Bold))
        charts_title.setAlignment(Qt.AlignLeft)
        charts_title.setStyleSheet(f"""
            QLabel {{
                color: #ffffff;
                padding: 12px 0px;
                background-color: transparent;
                border-radius: 0px;
                margin: 0px;
            }}
        """)
        charts_title_layout.addWidget(charts_title)
        
        charts_title_container = QWidget()
        charts_title_container.setStyleSheet(f"""
            QWidget {{
                background-color: {DUOC_PRIMARY};
                border-radius: 10px;
                margin: 10px 0px;
            }}
        """)
        charts_title_container.setLayout(charts_title_layout)
        charts_header.addWidget(charts_title_container)
        charts_header.addStretch()
        range_label = QLabel("Rango:")
        range_label.setFont(QFont("Arial", 11, QFont.Bold))
        charts_header.addWidget(range_label)
        self.range_combo = QComboBox()
        self.range_combo.addItems(["7 días", "14 días", "30 días"])
        self.range_combo.setCurrentIndex(0)
        self.range_combo.setStyleSheet("""
            QComboBox { padding: 6px 10px; border: 1px solid #dee2e6; border-radius: 6px; }
            QComboBox:focus { border-color: #003A70; }
        """)
        self.range_combo.currentTextChanged.connect(self.on_range_changed)
        charts_header.addWidget(self.range_combo)
        content_layout.addLayout(charts_header)
        
        # Contenedor principal para los gráficos en layout horizontal
        charts_container = QWidget()
        charts_container.setObjectName("ChartsContainer")
        charts_container.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Preferred)
        charts_container.setMaximumWidth(1400)
        charts_main_layout = QHBoxLayout(charts_container)
        charts_main_layout.setSpacing(15)
        charts_main_layout.setContentsMargins(5, 5, 5, 5)
        charts_main_layout.setAlignment(Qt.AlignCenter)
        
        # Gráfico 1: Visitantes por Día
        chart1_container = QFrame()
        chart1_layout = QVBoxLayout(chart1_container)
        chart1_layout.setSpacing(5)
        chart1_container.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e9ecef;
                border-radius: 10px;
                padding: 8px;
            }
        """)
        
        # Crear layout horizontal para título con icono
        chart1_title_layout = QHBoxLayout()
        chart1_title_layout.setContentsMargins(8, 0, 8, 0)
        chart1_title_layout.setSpacing(6)
        chart1_title_layout.setAlignment(Qt.AlignCenter)
        
        chart1_icon = get_icon_for_emoji("👥", 16)
        if not chart1_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(chart1_icon.pixmap(16, 16))
            chart1_title_layout.addWidget(icon_label)
        
        chart1_title = QLabel("Visitantes por Día")
        chart1_title.setFont(QFont("Arial", 14, QFont.Bold))
        chart1_title.setAlignment(Qt.AlignCenter)
        chart1_title.setStyleSheet("""
            QLabel {
                color: #ffffff;
                padding: 8px 0px;
                background-color: transparent;
                border-radius: 0px;
                margin: 0px;
            }
        """)
        chart1_title_layout.addWidget(chart1_title)
        
        chart1_title_container = QWidget()
        chart1_title_container.setStyleSheet(f"""
            QWidget {{
                background-color: {DUOC_PRIMARY};
                border-radius: 6px;
                margin: 2px;
            }}
        """)
        chart1_title_container.setLayout(chart1_title_layout)
        chart1_layout.addWidget(chart1_title_container)
        
        self.chart1 = ChartWidget()
        # Hacer el gráfico flexible dentro de su card
        self.chart1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        chart1_layout.addWidget(self.chart1)
        
        charts_main_layout.addWidget(chart1_container)
        
        # Gráfico 2: Estado de Visitantes
        chart2_container = QFrame()
        chart2_layout = QVBoxLayout(chart2_container)
        chart2_layout.setSpacing(5)
        chart2_container.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e9ecef;
                border-radius: 10px;
                padding: 8px;
            }
        """)
        
        # Crear layout horizontal para título con icono
        chart2_title_layout = QHBoxLayout()
        chart2_title_layout.setContentsMargins(8, 0, 8, 0)
        chart2_title_layout.setSpacing(6)
        chart2_title_layout.setAlignment(Qt.AlignCenter)
        
        chart2_icon = get_icon_for_emoji("👤", 16)
        if not chart2_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(chart2_icon.pixmap(16, 16))
            chart2_title_layout.addWidget(icon_label)
        
        chart2_title = QLabel("Estado de Visitantes")
        chart2_title.setFont(QFont("Arial", 14, QFont.Bold))
        chart2_title.setAlignment(Qt.AlignCenter)
        chart2_title.setStyleSheet("""
            QLabel {
                color: #000000;
                padding: 8px 0px;
                background-color: transparent;
                border-radius: 0px;
                margin: 0px;
            }
        """)
        chart2_title_layout.addWidget(chart2_title)
        
        chart2_title_container = QWidget()
        chart2_title_container.setStyleSheet(f"""
            QWidget {{
                background-color: {DUOC_SECONDARY};
                border-radius: 6px;
                margin: 2px;
            }}
        """)
        chart2_title_container.setLayout(chart2_title_layout)
        chart2_layout.addWidget(chart2_title_container)
        
        self.chart2 = ChartWidget()
        self.chart2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        chart2_layout.addWidget(self.chart2)
        
        charts_main_layout.addWidget(chart2_container)
        
        # Gráfico 3: Destinos Más Frecuentes
        chart3_container = QFrame()
        chart3_layout = QVBoxLayout(chart3_container)
        chart3_layout.setSpacing(5)
        chart3_container.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e9ecef;
                border-radius: 10px;
                padding: 8px;
            }
        """)
        
        # Crear layout horizontal para título con icono
        chart3_title_layout = QHBoxLayout()
        chart3_title_layout.setContentsMargins(8, 0, 8, 0)
        chart3_title_layout.setSpacing(6)
        chart3_title_layout.setAlignment(Qt.AlignCenter)
        
        chart3_icon = get_icon_for_emoji("🏢", 16)
        if not chart3_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(chart3_icon.pixmap(16, 16))
            chart3_title_layout.addWidget(icon_label)
        
        chart3_title = QLabel("Destinos Más Frecuentes")
        chart3_title.setFont(QFont("Arial", 14, QFont.Bold))
        chart3_title.setAlignment(Qt.AlignCenter)
        chart3_title.setStyleSheet("""
            QLabel {
                color: #000000;
                padding: 8px 0px;
                background-color: transparent;
                border-radius: 0px;
                margin: 0px;
            }
        """)
        chart3_title_layout.addWidget(chart3_title)
        
        chart3_title_container = QWidget()
        chart3_title_container.setStyleSheet(f"""
            QWidget {{
                background-color: {DUOC_SECONDARY};
                border-radius: 6px;
                margin: 2px;
            }}
        """)
        chart3_title_container.setLayout(chart3_title_layout)
        chart3_layout.addWidget(chart3_title_container)
        
        self.chart3 = ChartWidget()
        self.chart3.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        chart3_layout.addWidget(self.chart3)
        
        charts_main_layout.addWidget(chart3_container)

        # Ajustar proporciones más equilibradas (2:1:1) y misma altura mínima
        charts_main_layout.setStretch(0, 2)
        charts_main_layout.setStretch(1, 1)
        charts_main_layout.setStretch(2, 1)
        min_height = 360 if self.screen_config['value_font_size'] >= 24 else 320
        chart1_container.setMinimumHeight(min_height)
        chart2_container.setMinimumHeight(min_height)
        chart3_container.setMinimumHeight(min_height)
        
        # Agregar el contenedor de gráficos al layout principal centrado
        content_layout.addWidget(charts_container, alignment=Qt.AlignHCenter)
        
        
        # Tarjetas de métricas resumidas
        metrics_container = QWidget()
        metrics_layout = QHBoxLayout(metrics_container)
        metrics_layout.setSpacing(12)
        metrics_layout.setContentsMargins(0, 0, 0, 0)
        self.metric_cards = {}
        # Placeholders; se actualizan en update_statistics
        cards_spec = [
            ("👥", "Visitantes Hoy", "0", DUOC_SECONDARY, "Ingresos del día"),
            ("🏢", "Zonas Activas", "0", DUOC_PRIMARY, "Zonas con visitantes"),
            ("📈", "Total Visitas", "0", DUOC_PRIMARY, "Acumulado de registros"),
            ("⏱️", "Promedio Estancia", "N/A", DUOC_SECONDARY, "Tiempo promedio")
        ]
        for emoji, title, value, color, desc in cards_spec:
            card = self.create_stat_card(emoji, title, value, color, desc)
            self.metric_cards[title] = card
            metrics_layout.addWidget(card)
        content_layout.addWidget(metrics_container)

        # Reporte de Visitantes
        # Crear layout horizontal para título con icono
        current_visitors_title_layout = QHBoxLayout()
        current_visitors_title_layout.setContentsMargins(15, 0, 15, 0)
        current_visitors_title_layout.setSpacing(8)
        current_visitors_title_layout.setAlignment(Qt.AlignCenter)
        
        visitors_title_icon = get_icon_for_emoji("👥", 20)
        if not visitors_title_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(visitors_title_icon.pixmap(20, 20))
            current_visitors_title_layout.addWidget(icon_label)
        
        current_visitors_title = QLabel("Reporte de Visitantes")
        current_visitors_title.setFont(QFont("Arial", 18, QFont.Bold))
        current_visitors_title.setAlignment(Qt.AlignCenter)
        current_visitors_title.setStyleSheet("""
            QLabel {
                color: #2c3e50;
                padding: 15px 0px;
                background-color: transparent;
                border-radius: 0px;
                margin: 0px;
            }
        """)
        current_visitors_title_layout.addWidget(current_visitors_title)
        
        current_visitors_title_container = QWidget()
        current_visitors_title_container.setStyleSheet("""
            QWidget {
                background-color: #ecf0f1;
                border-radius: 10px;
                margin: 10px 0px;
            }
        """)
        current_visitors_title_container.setLayout(current_visitors_title_layout)
        content_layout.addWidget(current_visitors_title_container)
        
        # Botones de acción para el reporte
        action_layout = QHBoxLayout()
        
        # Botón para actualizar datos responsivo
        self.btn_refresh = QPushButton("Actualizar Datos")
        self.btn_refresh.setIcon(get_icon_for_emoji("🔄", 18))
        self.btn_refresh.setMinimumSize(self.screen_config['btn_width'], self.screen_config['btn_height'])
        self.btn_refresh.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.btn_refresh.setStyleSheet(get_standard_button_style(DUOC_SECONDARY))
        self.btn_refresh.clicked.connect(self.refresh_visitors_data)
        action_layout.addWidget(self.btn_refresh)
        
        # Filtro de visitantes responsivo
        filter_label = QLabel("Filtrar:")
        filter_label.setFont(QFont("Arial", self.screen_config['btn_font_size'], QFont.Bold))
        action_layout.addWidget(filter_label)
        
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Todos los visitantes", "Solo visitantes actuales", "Solo visitantes que se fueron"])
        self.filter_combo.setMinimumSize(self.screen_config['btn_width'] * 1.3, self.screen_config['btn_height'])
        self.filter_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.filter_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 5px;
                border: 2px solid {duoc_lighten(DUOC_PRIMARY, 0.6)};
                border-radius: 5px;
                font-size: {self.screen_config['btn_font_size']}px;
            }}
            QComboBox:focus {{
                border-color: {DUOC_PRIMARY};
            }}
        """)
        self.filter_combo.currentTextChanged.connect(self.refresh_visitors_data)
        action_layout.addWidget(self.filter_combo)
        
        # Botón para exportar a Excel responsivo
        self.btn_export = QPushButton("📊 Exportar a Excel")
        self.btn_export.setMinimumSize(self.screen_config['btn_width'], self.screen_config['btn_height'])
        self.btn_export.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.btn_export.setStyleSheet(get_standard_button_style(DUOC_PRIMARY))
        self.btn_export.clicked.connect(self.export_to_excel)
        action_layout.addWidget(self.btn_export)
        
        action_layout.addStretch()
        content_layout.addLayout(action_layout)
        
        # Tabla de visitantes
        self.visitors_table = QTableWidget()
        self.visitors_table.setColumnCount(8)  # 8 columnas incluyendo usuario registrador
        from PySide6.QtWidgets import QTableWidgetItem
        header_icons = ["👤", "📄", "📅", "📅", "🏢", "🤝", "📍", "👨‍💼"]
        header_labels = ["Nombre", "RUT", "Entrada", "Salida", "Destino", "Acompañante", "Estado", "Registrado por"]
        for i, (icon_emoji, label) in enumerate(zip(header_icons, header_labels)):
            header_item = QTableWidgetItem(label)
            icon = get_icon_for_emoji(icon_emoji, 16)
            if not icon.isNull():
                header_item.setIcon(icon)
            self.visitors_table.setHorizontalHeaderItem(i, header_item)

        configure_modern_table(self.visitors_table, row_height=self.screen_config['row_height'])
        apply_modern_table_theme(self.visitors_table)

        # Scrollbars según necesidad
        self.visitors_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.visitors_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Configurar distribución de columnas para mejor legibilidad
        header = self.visitors_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)          # Nombre (expandible)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents) # RUT (compacto)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents) # Fecha entrada (compacto)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents) # Fecha salida (compacto)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents) # Destino (compacto)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents) # Acompañante (compacto)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents) # Estado (compacto)
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents) # Usuario registrador (compacto)

        # Establecer tamaños mínimos para mejor legibilidad
        header.setMinimumSectionSize(100)  # Tamaño mínimo para todas las columnas
        header.setDefaultSectionSize(140)  # Tamaño por defecto más generoso

        # Configurar la tabla para que use todo el espacio disponible
        self.visitors_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        content_layout.addWidget(self.visitors_table, 1)  # El 1 hace que tome todo el espacio disponible

        # Estado vacío
        self.empty_state = QLabel("\nNo hay datos para mostrar con el filtro seleccionado.\n")
        self.empty_state.setAlignment(Qt.AlignCenter)
        self.empty_state.setStyleSheet("color: #6c757d; border: 1px dashed #dee2e6; border-radius: 10px; padding: 16px;")
        self.empty_state.setVisible(False)
        content_layout.addWidget(self.empty_state)
        
        # Información del reporte
        self.info_label = QLabel()
        self.info_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.info_label.setStyleSheet("""
            QLabel {
                color: #2c3e50;
                padding: 15px;
                background-color: #e8f4fd;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin: 10px 0px;
            }
        """)
        content_layout.addWidget(self.info_label)
        
        # Cargar datos iniciales
        self.refresh_visitors_data()
        self.update_statistics()
        
        # Configurar el área de scroll con el contenido
        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area, 1)  # El 1 hace que tome todo el espacio disponible
        
        # Botón de regreso
        back_button = QPushButton("Volver al Menú Principal")
        back_button.setIcon(get_icon_for_emoji("⬅️", 18))
        back_button.setFixedSize(200, 40)
        back_button.setStyleSheet(
            f"""
            QPushButton {{
                background-color: rgba(255, 184, 28, 0.1);
                color: {DUOC_SECONDARY};
                border: 2px solid {DUOC_SECONDARY};
                border-radius: 14px;
                padding: 10px 20px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: {DUOC_SECONDARY};
                color: #000000;
            }}
            """
        )
        back_button.clicked.connect(self.go_to_main)
        main_layout.addWidget(back_button, alignment=Qt.AlignCenter)
    
    def go_to_main(self):
        """Regresa al menú principal"""
        # Buscar la ventana principal que contiene el navigation_manager
        parent = self.parent()
        while parent is not None:
            if hasattr(parent, 'navigation_manager'):
                parent.navigation_manager.navigate_to("main")
                return
            elif hasattr(parent, 'go_to_main'):
                parent.go_to_main()
                return
            parent = parent.parent()
    
    def create_stat_card(self, emoji, title, value, color, description):
        """Crea una tarjeta de estadística responsiva"""
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel)
        
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: {self.screen_config['border_radius']}px;
                padding: {self.screen_config['padding']}px;
            }}
            QFrame:hover {{
                background-color: {self.lighten_color(color)};
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        
        # Valor
        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", self.screen_config['value_font_size'], QFont.Bold))
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet(f"color: {color};")
        layout.addWidget(value_label)
        
        # Título con icono
        title_layout = QHBoxLayout()
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.setSpacing(6)
        title_layout.setAlignment(Qt.AlignCenter)
        
        title_icon = get_icon_for_emoji(emoji, 18)
        if not title_icon.isNull():
            icon_label = QLabel()
            icon_label.setPixmap(title_icon.pixmap(18, 18))
            title_layout.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", self.screen_config['title_font_size'], QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #000000;")
        title_layout.addWidget(title_label)
        
        title_container = QWidget()
        title_container.setLayout(title_layout)
        layout.addWidget(title_container)
        
        # Descripción
        desc_label = QLabel(description)
        desc_label.setFont(QFont("Arial", self.screen_config['desc_font_size']))
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("color: #6c757d;")
        layout.addWidget(desc_label)
        
        return card
    
    def create_stat_card_with_reference(self, title, value, color, description):
        """Crea una tarjeta de estadística responsiva y devuelve la tarjeta y el label de valor"""
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel)
        
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: {self.screen_config['border_radius']}px;
                padding: {self.screen_config['padding']}px;
            }}
            QFrame:hover {{
                background-color: {self.lighten_color(color)};
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        
        # Valor (con referencia)
        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", self.screen_config['value_font_size'], QFont.Bold))
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet(f"color: {color};")
        layout.addWidget(value_label)
        
        # Título
        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", self.screen_config['title_font_size'], QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        title_label.setStyleSheet("color: #000000;")
        
        # Descripción
        desc_label = QLabel(description)
        desc_label.setFont(QFont("Arial", self.screen_config['desc_font_size']))
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("color: #6c757d;")
        layout.addWidget(desc_label)
        
        return card, value_label
    
    def create_report_card(self, title, color, description):
        """Crea una tarjeta de reporte"""
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 12px;
                padding: 15px;
            }}
            QFrame:hover {{
                background-color: {self.lighten_color(color)};
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(10)
        
        # Título
        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"color: {color};")
        layout.addWidget(title_label)
        title_label.setStyleSheet("color: #000000;")
        
        # Descripción
        desc_label = QLabel(description)
        desc_label.setFont(QFont("Arial", 10))
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("color: #6c757d;")
        layout.addWidget(desc_label)
        
        # Botón de generar
        generate_btn = QPushButton("📄 Generar")
        generate_btn.setFixedHeight(35)
        generate_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
        """)
        layout.addWidget(generate_btn)
        
        return card
    
    def refresh_visitors_data(self):
        """Actualiza los datos de visitantes en la tabla"""
        try:
            # Recargar los datos desde el archivo JSON para obtener la información más actualizada
            self.visitor_manager.load_visitors()
            
            # Determinar qué datos mostrar según el filtro
            filter_text = self.filter_combo.currentText()
            if filter_text == "Solo visitantes actuales":
                visitors_data = self.visitor_manager.get_visitor_report_data(include_departed=False)
            elif filter_text == "Solo visitantes que se fueron":
                # Filtrar solo los que se fueron
                all_data = self.visitor_manager.get_visitor_report_data(include_departed=True)
                visitors_data = [v for v in all_data if v['estado_visita'] == "Finalizada"]
            else:  # "Todos los visitantes"
                visitors_data = self.visitor_manager.get_visitor_report_data(include_departed=True)
            
            # Configurar número de filas
            self.visitors_table.setRowCount(len(visitors_data))
            
            # Llenar la tabla con los datos
            for row, visitor in enumerate(visitors_data):
                # Crear items con texto centrado y mejor formato
                # Mostrar RUT normalizado
                rut_display = format_rut_display(visitor['rut']) if visitor['rut'] else "N/A"
                
                items = [
                    QTableWidgetItem(str(visitor['nombre'])),
                    QTableWidgetItem(rut_display),
                    QTableWidgetItem(str(visitor['fecha_entrada'])),
                    QTableWidgetItem(str(visitor['fecha_salida'])),
                    QTableWidgetItem(str(visitor['destino'])),
                    QTableWidgetItem(str(visitor['acompañante'])),
                    QTableWidgetItem(str(visitor['estado_visita'])),
                    QTableWidgetItem(str(visitor.get('usuario_registrador', 'Sistema')))
                ]
                
                # Configurar alineación y estilo para cada item usando configuración centralizada
                for col, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    item.setFont(QFont("Arial", self.screen_config['font_size']))
                    
                    # Colorear según el estado
                    if col == 3:  # Fecha de salida
                        if visitor['fecha_salida'] == "Aún en el edificio":
                            item.setBackground(QColor(220, 248, 198))  # Verde claro
                            item.setForeground(QColor(40, 167, 69))   # Verde oscuro
                        else:
                            item.setBackground(QColor(248, 249, 250))  # Gris claro
                            item.setForeground(QColor(108, 117, 125)) # Gris oscuro
                    
                    elif col == 6:  # Estado de visita
                        if visitor['estado_visita'] == "En curso":
                            item.setBackground(QColor(220, 248, 198))  # Verde claro
                            item.setForeground(QColor(40, 167, 69))   # Verde oscuro
                        else:  # Finalizada
                            item.setBackground(QColor(248, 215, 218))  # Rojo claro
                            item.setForeground(QColor(220, 53, 69))   # Rojo oscuro
                    
                    self.visitors_table.setItem(row, col, item)
            
            # Ajustar altura de filas
            for row in range(len(visitors_data)):
                self.visitors_table.setRowHeight(row, self.screen_config['row_height'])
            
            # Las esquinas inferiores se manejan con CSS
            
            # Alternar estado vacío
            no_rows = len(visitors_data) == 0
            self.empty_state.setVisible(no_rows)
            self.visitors_table.setVisible(not no_rows)

            # Actualizar información del reporte
            self.update_report_info(visitors_data)
            
            # Actualizar estadísticas
            self.update_statistics()
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error al cargar datos: {str(e)}")
    
    def update_report_info(self, visitors_data):
        """Actualiza la información del reporte"""
        total_visitors = len(visitors_data)
        
        # Contar visitantes por destino
        destinations = {}
        for visitor in visitors_data:
            dest = visitor['destino']
            destinations[dest] = destinations.get(dest, 0) + 1
        
        # Crear texto informativo
        info_text = f"📊 Total de visitantes actuales: {total_visitors}"
        if destinations:
            dest_info = ", ".join([f"{dest}: {count}" for dest, count in destinations.items()])
            info_text += f" | 📍 Distribución por destino: {dest_info}"
        
        self.info_label.setText(info_text)
    
    def update_statistics(self):
        """Actualiza los gráficos con datos actuales"""
        try:
            # Obtener todos los visitantes
            all_visitors = self.visitor_manager.get_all_visitors()
            print(f"Debug: Encontrados {len(all_visitors)} visitantes")
            
            # Actualizar los gráficos
            self.update_charts(all_visitors)

            # Actualizar tarjetas de métricas
            stats = self.calculate_statistics(all_visitors)
            cards_map = {
                "Visitantes Hoy": stats.get('visitors_today', '0'),
                "Zonas Activas": stats.get('active_zones', '0'),
                "Total Visitas": stats.get('total_visits', '0'),
                "Promedio Estancia": stats.get('avg_visit_time', 'N/A'),
            }
            for title, value in cards_map.items():
                card = self.metric_cards.get(title)
                if card:
                    # actualizar el primer label (valor) dentro del card
                    layout = card.layout()
                    if layout and layout.count() > 0:
                        w = layout.itemAt(0).widget()
                        if isinstance(w, QLabel):
                            w.setText(str(value))
            
            print("Debug: Gráficos actualizados en la interfaz")
            
        except Exception as e:
            print(f"Error al actualizar gráficos: {e}")
            import traceback
            traceback.print_exc()
    
    def update_charts(self, visitors):
        """Actualiza los gráficos con datos actuales"""
        try:
            # Calcular datos para los gráficos
            visitors_by_day = self.calculate_visitors_by_day(visitors)
            current_visitors, departed_visitors = self.calculate_visitor_status(visitors)
            popular_destinations = self.calculate_popular_destinations(visitors)
            
            # Actualizar los gráficos
            items = list(visitors_by_day.items())
            limit = max(1, min(self.range_days, len(items)))
            limited_visitors_by_day = dict(items[-limit:])
            self.chart1.plot_visitors_by_day(limited_visitors_by_day)
            self.chart2.plot_visitors_status(current_visitors, departed_visitors)
            self.chart3.plot_popular_destinations(popular_destinations)
            
            print("Debug: Gráficos actualizados correctamente")
            
        except Exception as e:
            print(f"Error al actualizar gráficos: {e}")
            import traceback
            traceback.print_exc()

    def on_range_changed(self, text: str):
        """Cambia el rango del gráfico 1 sin alterar la lógica de datos."""
        try:
            if text.startswith("7"):
                self.range_days = 7
            elif text.startswith("14"):
                self.range_days = 14
            else:
                self.range_days = 30
        except Exception:
            self.range_days = 7
        # Redibujar con el nuevo rango
        self.update_statistics()
    
    def calculate_visitors_by_day(self, visitors):
        """Calcula visitantes por día para el gráfico"""
        from datetime import datetime, timedelta
        
        # Obtener los últimos 7 días
        today = datetime.now().date()
        visitors_by_day = {}
        
        for i in range(7):
            date = today - timedelta(days=i)
            visitors_by_day[date] = 0
        
        # Contar visitantes por día
        for visitor in visitors:
            try:
                visit_date = datetime.strptime(visitor.fecha_ingreso, "%Y-%m-%d %H:%M:%S").date()
                if visit_date in visitors_by_day:
                    visitors_by_day[visit_date] += 1
            except:
                continue
        
        # Ordenar por fecha
        sorted_dates = sorted(visitors_by_day.keys())
        return {date: visitors_by_day[date] for date in sorted_dates}
    
    def calculate_visitor_status(self, visitors):
        """Calcula visitantes actuales vs que se fueron"""
        current = len([v for v in visitors if v.estado == "Dentro"])
        departed = len([v for v in visitors if v.estado == "Fuera"])
        return current, departed
    
    def calculate_popular_destinations(self, visitors):
        """Calcula los destinos más populares"""
        destinations = {}
        for visitor in visitors:
            dest = visitor.sector
            destinations[dest] = destinations.get(dest, 0) + 1
        return destinations
    
    def calculate_statistics(self, visitors):
        """Calcula las estadísticas basadas en los datos de visitantes"""
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        current_time = datetime.now()
        
        # Visitantes de hoy
        visitors_today = 0
        for visitor in visitors:
            try:
                visit_date = datetime.strptime(visitor.fecha_ingreso, "%Y-%m-%d %H:%M:%S").date()
                if visit_date == today:
                    visitors_today += 1
            except:
                continue
        
        # Zonas activas (con visitantes actuales)
        active_zones = set()
        for visitor in visitors:
            if visitor.estado == "Dentro":
                active_zones.add(visitor.sector)
        
        # Total de visitas
        total_visits = len(visitors)
        
        # Tiempo promedio de visita
        avg_visit_time = self.calculate_average_visit_time(visitors)
        
        # Estadísticas adicionales
        most_visited_zone = self.get_most_visited_zone(visitors)
        current_visitors = len([v for v in visitors if v.estado == "Dentro"])
        visits_this_week = self.get_visits_this_week(visitors)
        longest_visit = self.get_longest_visit(visitors)
        
        return {
            'visitors_today': str(visitors_today),
            'active_zones': str(len(active_zones)),
            'total_visits': str(total_visits),
            'avg_visit_time': avg_visit_time,
            'most_visited_zone': most_visited_zone,
            'current_visitors': str(current_visitors),
            'visits_this_week': str(visits_this_week),
            'longest_visit': longest_visit
        }
    
    def calculate_average_visit_time(self, visitors):
        """Calcula el tiempo promedio de visita en minutos"""
        completed_visits = []
        
        for visitor in visitors:
            if visitor.fecha_salida:
                try:
                    entry_time = datetime.strptime(visitor.fecha_ingreso, "%Y-%m-%d %H:%M:%S")
                    exit_time = datetime.strptime(visitor.fecha_salida, "%Y-%m-%d %H:%M:%S")
                    duration = (exit_time - entry_time).total_seconds() / 60  # en minutos
                    completed_visits.append(duration)
                except:
                    continue
        
        if completed_visits:
            avg_minutes = sum(completed_visits) / len(completed_visits)
            if avg_minutes < 60:
                return f"{avg_minutes:.0f} min"
            else:
                hours = avg_minutes / 60
                return f"{hours:.1f} h"
        else:
            return "N/A"
    
    def get_most_visited_zone(self, visitors):
        """Obtiene la zona más visitada"""
        zone_counts = {}
        for visitor in visitors:
            zone = visitor.sector
            zone_counts[zone] = zone_counts.get(zone, 0) + 1
        
        if zone_counts:
            most_visited = max(zone_counts, key=zone_counts.get)
            count = zone_counts[most_visited]
            return f"{most_visited} ({count})"
        else:
            return "N/A"
    
    def get_visits_this_week(self, visitors):
        """Calcula las visitas de esta semana"""
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        
        visits_this_week = 0
        for visitor in visitors:
            try:
                visit_date = datetime.strptime(visitor.fecha_ingreso, "%Y-%m-%d %H:%M:%S").date()
                if visit_date >= week_ago:
                    visits_this_week += 1
            except:
                continue
        
        return visits_this_week
    
    def get_longest_visit(self, visitors):
        """Obtiene la duración de la visita más larga"""
        longest_duration = 0
        
        for visitor in visitors:
            if visitor.fecha_salida:
                try:
                    entry_time = datetime.strptime(visitor.fecha_ingreso, "%Y-%m-%d %H:%M:%S")
                    exit_time = datetime.strptime(visitor.fecha_salida, "%Y-%m-%d %H:%M:%S")
                    duration = (exit_time - entry_time).total_seconds() / 60  # en minutos
                    longest_duration = max(longest_duration, duration)
                except:
                    continue
        
        if longest_duration > 0:
            if longest_duration < 60:
                return f"{longest_duration:.0f} min"
            else:
                hours = longest_duration / 60
                return f"{hours:.1f} h"
        else:
            return "N/A"
    
    def update_stat_card(self, card, value):
        """Actualiza el valor de una tarjeta de estadística"""
        # Buscar el label de valor en la tarjeta (es el primer QLabel con font bold)
        layout = card.layout()
        for i in range(layout.count()):
            item = layout.itemAt(i)
            if item and item.widget():
                widget = item.widget()
                if isinstance(widget, QLabel) and widget.font().bold():
                    # Verificar que sea el label de valor (no el título)
                    if widget.text() != "0" and not any(keyword in widget.text() for keyword in ["Visitantes", "Zonas", "Visitas", "Promedio", "Zona Más", "Actuales", "Semana", "Larga"]):
                        widget.setText(str(value))
                        print(f"Debug: Actualizando tarjeta con valor: {value}")
                        return
                    elif widget.text() == "0" or widget.text() == "N/A":
                        widget.setText(str(value))
                        print(f"Debug: Actualizando tarjeta con valor: {value}")
                        return
        print(f"Debug: No se encontró el label para actualizar en la tarjeta")
    
    def export_to_excel(self):
        """Exporta los datos de visitantes a Excel (solo administradores)"""
        # Verificar permisos de administrador
        if not self.auth_manager.is_admin():
            QMessageBox.warning(
                self, 
                "Acceso Denegado", 
                "Solo los administradores pueden exportar reportes a Excel."
            )
            return
        
        try:
            # Recargar los datos desde el archivo JSON para obtener la información más actualizada
            self.visitor_manager.load_visitors()
            
            # Determinar qué datos exportar según el filtro
            filter_text = self.filter_combo.currentText()
            if filter_text == "Solo visitantes actuales":
                visitors_data = self.visitor_manager.get_visitor_report_data(include_departed=False)
            elif filter_text == "Solo visitantes que se fueron":
                # Filtrar solo los que se fueron
                all_data = self.visitor_manager.get_visitor_report_data(include_departed=True)
                visitors_data = [v for v in all_data if v['estado_visita'] == "Finalizada"]
            else:  # "Todos los visitantes"
                visitors_data = self.visitor_manager.get_visitor_report_data(include_departed=True)
            
            if not visitors_data:
                QMessageBox.information(self, "Información", f"No hay visitantes para exportar con el filtro '{filter_text}'.")
                return
            
            # Solicitar ubicación del archivo
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte de Visitantes",
                f"reporte_visitantes_actuales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Archivos Excel (*.xlsx);;Todos los archivos (*)"
            )
            
            if filename:
                # Crear DataFrame con los datos
                import pandas as pd
                df = pd.DataFrame(visitors_data)
                
                # Asegurar que tenemos todas las columnas necesarias incluyendo usuario_registrador
                expected_columns = [
                    'nombre', 'rut', 'fecha_entrada', 'fecha_salida', 
                    'destino', 'acompañante', 'estado_visita', 'usuario_registrador'
                ]
                
                # Filtrar solo las columnas que necesitamos
                df = df[expected_columns]
                
                # Renombrar columnas para mejor presentación
                df.columns = [
                    'Nombre Completo',
                    'RUN',
                    'Fecha y Hora de Entrada',
                    'Fecha y Hora de Salida',
                    'Destino',
                    'Acompañante',
                    'Estado de Visita',
                    'Registrado por'
                ]
                
                # Crear archivo Excel con formato
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Visitantes Actuales', index=False)
                    
                    # Obtener el workbook y worksheet para aplicar formato
                    workbook = writer.book
                    worksheet = writer.sheets['Visitantes Actuales']
                    
                    # Aplicar formato a las columnas
                    self._format_excel_worksheet(worksheet, len(visitors_data))
                
                # Mostrar mensaje de éxito
                QMessageBox.information(
                    self, 
                    "Exportación Exitosa", 
                    f"El reporte se ha guardado exitosamente en:\n{filename}"
                )
                
                # Preguntar si desea borrar los datos exportados
                reply = QMessageBox.question(
                    self,
                    "Borrar Datos Exportados",
                    "¿Desea borrar los datos exportados de la base de datos?\n\n"
                    "⚠️ ADVERTENCIA: Esta acción NO se puede deshacer.\n"
                    "Todos los registros de visitantes serán eliminados permanentemente.",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    # Confirmar nuevamente para evitar borrados accidentales
                    confirm = QMessageBox.warning(
                        self,
                        "Confirmar Eliminación",
                        "¿Está COMPLETAMENTE SEGURO de que desea borrar TODOS los datos?\n\n"
                        "Esta es su última oportunidad para cancelar.",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    
                    if confirm == QMessageBox.Yes:
                        # Borrar todos los visitantes
                        if self.visitor_manager.delete_all_visitors():
                            QMessageBox.information(
                                self,
                                "Datos Eliminados",
                                "✅ Todos los datos de visitantes han sido eliminados exitosamente de la base de datos."
                            )
                            # Actualizar la interfaz
                            self.refresh_visitors_data()
                        else:
                            QMessageBox.critical(
                                self,
                                "Error",
                                "❌ Hubo un error al eliminar los datos. Por favor, intente nuevamente."
                            )
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar a Excel: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _format_excel_worksheet(self, worksheet, num_rows):
        """Aplica formato al worksheet de Excel"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Estilo para el encabezado
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Estilo para las celdas de datos
            data_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Limpiar cualquier formato previo que pueda causar íconos
            for row in range(1, num_rows + 2):
                for col in range(1, 9):  # 8 columnas incluyendo usuario registrador
                    cell = worksheet.cell(row=row, column=col)
                    # Limpiar comentarios y validaciones que puedan causar íconos
                    if cell.comment:
                        cell.comment = None
                    if hasattr(cell, 'data_validation') and cell.data_validation:
                        cell.data_validation = None
            
            # Aplicar formato al encabezado (fila 1)
            for col in range(1, 9):  # 8 columnas incluyendo usuario registrador
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Aplicar formato a las celdas de datos
            for row in range(2, num_rows + 2):  # +2 porque empezamos desde la fila 2
                for col in range(1, 9):  # 8 columnas incluyendo usuario registrador
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Ajustar ancho de columnas
            column_widths = [25, 15, 20, 20, 20, 20, 15, 18]  # Anchos para cada columna (incluyendo usuario registrador)
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = width
                
        except Exception as e:
            print(f"Error al aplicar formato al Excel: {e}")
    
    def lighten_color(self, color, factor=0.1):
        """Aclara un color hexadecimal - usa la función del tema si está disponible"""
        try:
            return duoc_lighten(color, factor)
        except:
            # Fallback si no está disponible la función del tema
            color = color.lstrip('#')
            r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            r = min(255, int(r + (255 - r) * factor))
            g = min(255, int(g + (255 - g) * factor))
            b = min(255, int(b + (255 - b) * factor))
            return f"#{r:02x}{g:02x}{b:02x}"
    
    def darken_color(self, color, factor=0.2):
        """Oscurece un color hexadecimal - usa la función del tema si está disponible"""
        try:
            return duoc_darken(color, factor)
        except:
            # Fallback si no está disponible la función del tema
            color = color.lstrip('#')
            r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            r = max(0, int(r * (1 - factor)))
            g = max(0, int(g * (1 - factor)))
            b = max(0, int(b * (1 - factor)))
            return f"#{r:02x}{g:02x}{b:02x}"
